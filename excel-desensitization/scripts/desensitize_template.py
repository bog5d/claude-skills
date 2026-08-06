#!/usr/bin/env python3
"""
加密 Excel → 解密 → 脱敏 → 加密 完整模板
用法：修改下方 COMPANY_REPLACEMENTS 映射表和密码后直接运行
"""

import xlrd
import openpyxl
from openpyxl.styles import Font
import random
import re
import sys

random.seed(42)  # 可复现

# ═══════════════ 配置区 ═══════════════
SRC = "decrypted_input.xls"          # 解密后的文件
DST = "desensitized_output.xlsx"     # 脱敏输出（中间文件）
ENC_DST = "desensitized_output_加密.xlsx"  # 最终加密文件
PASSWORD = "your_password_here"

COMPANY_REPLACEMENTS = {
    # 主公司
    '原始公司全名': 'XX科技有限公司',
    # 子公司
    '原始子公司A全名': '子公司A',
    '原始子公司B全名': '子公司B',
}
# ═══════════════════════════════════

client_map = {}
supplier_map = {}

def get_client_name(original):
    if not original or not original.strip():
        return original
    o = original.strip()
    if o not in client_map:
        client_map[o] = f'客户{len(client_map)+1}'
    return client_map[o]

def get_supplier_name(original):
    if not original or not original.strip():
        return original
    o = original.strip()
    if o not in supplier_map:
        supplier_map[o] = f'供应商{len(supplier_map)+1}'
    return supplier_map[o]

def looks_like_amount(val):
    """判断是否像金额（排除年份/序号/日期序列等）"""
    if isinstance(val, float):
        if 2000 <= val <= 2099 and val == int(val):
            return False
        # Excel 日期序列号 (1982~2064)
        if 30000 <= val <= 60000 and val == int(val):
            return False
        if val == int(val) and abs(val) <= 100:
            return False
        return abs(val) > 1 or (abs(val) > 0 and '.' in str(val))
    return False

def perturb_amount(val):
    """随机扰动金额 ±30%"""
    factor = random.uniform(0.70, 1.30)
    new_val = val * factor
    if val == int(val):
        return round(new_val, 0)
    s = str(val)
    if '.' in s:
        decimals = len(s.split('.')[1])
        return round(new_val, decimals)
    return round(new_val, 2)

def desensitize_cell(cell_value, sheet_name, col_idx):
    """对单个单元格值脱敏"""
    s = str(cell_value).strip()

    # 1. 公司名替换
    for old, new in COMPANY_REPLACEMENTS.items():
        if old in s:
            s = s.replace(old, new)

    # 2. 客户/供应商名替换（根据Sheet名判断上下文）
    if '往来款项' in sheet_name or '应收' in sheet_name or '应付' in sheet_name:
        if col_idx == 2 and s and s not in ['', '往来单位名称', '客户名称', '供应商名称'] and not s.startswith('23'):
            s = get_client_name(s)
    elif '销售' in sheet_name or '采购' in sheet_name:
        if col_idx == 2 and s and s not in ['', '客户名称', '供应商名称'] and not s.startswith('23'):
            s = get_client_name(s)

    # 3. 金额处理
    if isinstance(cell_value, float) and looks_like_amount(cell_value):
        return perturb_amount(cell_value)

    # 4. 签署人替换
    s = re.sub(r'法定代表人[：:]?\s*\S+', '法定代表人：张XX', s)
    s = re.sub(r'主管会计工作负责人[：:]?\s*\S+', '主管会计工作负责人：李XX', s)
    s = re.sub(r'会计机构负责人[：:]?\s*\S+', '会计机构负责人：王XX', s)

    return s


def main():
    # ── 读取原始 xls ──
    wb_old = xlrd.open_workbook(SRC, formatting_info=True)

    # ── 创建新 xlsx ──
    wb_new = openpyxl.Workbook()
    wb_new.remove(wb_new.active)

    normal_font = Font(name='微软雅黑', size=10)

    for sn in wb_old.sheet_names():
        new_sn = sn
        for old, new in COMPANY_REPLACEMENTS.items():
            new_sn = new_sn.replace(old, new)
        if len(new_sn) > 31:
            new_sn = new_sn[:31]

        ws = wb_new.create_sheet(title=new_sn)
        sh = wb_old.sheet_by_name(sn)

        for r in range(sh.nrows):
            for c in range(sh.ncols):
                cell_value = sh.cell_value(r, c)
                new_val = desensitize_cell(cell_value, sn, c)
                cell = ws.cell(row=r+1, column=c+1)
                cell.value = new_val
                cell.font = normal_font

        ws.freeze_panes = 'A2'

        # 自动列宽
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    val_str = str(cell.value)
                    length = sum(2 if ord(ch) > 127 else 1 for ch in val_str)
                    max_length = max(max_length, length)
            ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    print(f"脱敏完成，共处理 {len(wb_old.sheet_names())} 个Sheet")
    print(f"客户映射: {len(client_map)} 个")
    print(f"供应商映射: {len(supplier_map)} 个")

    wb_new.save(DST)
    print(f"已保存脱敏文件: {DST}")

    # ── 加密 ──
    import msoffcrypto
    with open(DST, 'rb') as f:
        of = msoffcrypto.OfficeFile(f)
        with open(ENC_DST, 'wb') as outf:
            of.encrypt(PASSWORD, outf)
    print(f"已加密: {ENC_DST}")


if __name__ == '__main__':
    main()
