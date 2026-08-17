# -*- coding: utf-8 -*-
"""爆雷台账 -> xlsx（状态列可勾选、优先级配色、冻结首行）
用法：把 rows 列表换成你的台账数据后 python3 gen_ledger_xlsx.py
列：编号/雷型/优先级/原话摘录/场次/雷点/修法/状态/责任人建议
状态值：🔴未改 / 🟡部分改 / ✅已改
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = os.path.expanduser("~/录音拆解/analysis/01_全量爆雷台账.xlsx")  # 改输出路径

rows = [
    # (编号, 雷型, 优先级, 原话摘录, 场次, 雷点, 修法, 状态, 责任人建议)
    ("A1", "37.8亿订单化", "P0", "预测的并单在37.8亿元以上…可以拿到的订单", "臣易",
     "机会清单冒充在手订单，尽调砍可信度", "标准句：十五五机会清单+6月末在手约X", "🔴未改", "宋佳莹/市场部"),
    # ... 全部条目
]

wb = Workbook()
ws = wb.active
ws.title = "爆雷台账"

headers = ["编号", "雷型", "优先级", "原话摘录", "场次", "雷点", "修法", "状态", "责任人建议"]
ws.append(headers)

head_fill = PatternFill("solid", fgColor="1F3B66")
head_font = Font(bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
p0_fill = PatternFill("solid", fgColor="FDE9E9")
p1_fill = PatternFill("solid", fgColor="FFF7E6")
red_font = Font(color="C00000", bold=True)

for c in ws[1]:
    c.fill = head_fill; c.font = head_font; c.border = border
    c.alignment = Alignment(horizontal="center", vertical="center")

for r in rows:
    ws.append(list(r))

for row in ws.iter_rows(min_row=2):
    for c in row:
        c.border = border
        c.alignment = Alignment(vertical="top", wrap_text=True)
    p = row[2].value
    if p == "P0":
        for c in row:
            c.fill = p0_fill
    elif p == "P1":
        for c in row:
            c.fill = p1_fill
    if row[7].value == "🔴未改":
        row[7].font = red_font

widths = [7, 13, 8, 44, 14, 34, 38, 10, 14]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + i)].width = w
ws.freeze_panes = "A2"

wb.save(OUT)
print("OK:", OUT, f"{len(rows)} 条")
